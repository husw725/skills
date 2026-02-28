#!/usr/bin/env python3
import argparse
import csv
import io
import json
import os
import re
import sys
from datetime import datetime, date
from typing import Any, Dict, List, Tuple, Optional


def eprint(*a, **k):
    print(*a, file=sys.stderr, **k)


def parse_args():
    p = argparse.ArgumentParser(description="Write/append/replace_sheet in an Excel .xlsx from JSON/CSV/Markdown table")
    p.add_argument("--out", required=True, help="Output xlsx path")
    p.add_argument("--sheet", default="Sheet1", help="Sheet name")
    p.add_argument("--mode", choices=["overwrite", "append", "replace_sheet"], default="overwrite")
    p.add_argument("--input", default="-", help="Input file path or '-' for stdin")
    p.add_argument("--format", choices=["json", "csv", "md"], required=True)
    return p.parse_args()


def read_text(path: str) -> str:
    if path == "-":
        return sys.stdin.read()
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def parse_json(text: str) -> List[Dict[str, Any]]:
    data = json.loads(text)
    if isinstance(data, dict):
        # allow {"rows": [...]} convenience
        if "rows" in data and isinstance(data["rows"], list):
            data = data["rows"]
    if not isinstance(data, list):
        raise ValueError("JSON input must be an array of objects")
    rows: List[Dict[str, Any]] = []
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"Row {i} is not an object")
        rows.append(item)
    return rows


def parse_csv_text(text: str) -> Tuple[List[str], List[List[str]]]:
    # sniff delimiter lightly
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except Exception:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    all_rows = [r for r in reader if any(c.strip() for c in r)]
    if not all_rows:
        return [], []
    header = all_rows[0]
    body = all_rows[1:]
    return header, body


def parse_md_table(text: str) -> Tuple[List[str], List[List[str]]]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    # find first table block
    table = []
    for ln in lines:
        if ln.startswith("|") and ln.endswith("|"):
            table.append(ln)
        elif table:
            break
    if len(table) < 2:
        raise ValueError("No markdown table found")
    # header
    def split_row(ln: str) -> List[str]:
        parts = [p.strip() for p in ln.strip("|").split("|")]
        return parts
    header = split_row(table[0])
    # skip alignment row (---)
    body = []
    for ln in table[2:]:
        body.append(split_row(ln))
    return header, body


def union_keys_order(rows: List[Dict[str, Any]]) -> List[str]:
    cols: List[str] = []
    seen = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                cols.append(k)
    return cols


def coerce_cell(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float, bool, datetime, date)):
        return v
    if isinstance(v, str):
        s = v.strip()
        # number
        if re.fullmatch(r"[-+]?\d+", s):
            try:
                return int(s)
            except Exception:
                return v
        if re.fullmatch(r"[-+]?\d*\.\d+", s):
            try:
                return float(s)
            except Exception:
                return v
        # iso datetime/date
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.date() if fmt in ("%Y-%m-%d", "%Y/%m/%d") else dt
            except Exception:
                pass
        return v
    # fallback: stringify complex
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception as ex:
        raise RuntimeError(
            "Missing dependency 'openpyxl'. Install with: python3 -m pip install --user openpyxl"
        ) from ex


def autofit_columns(ws, max_width: int = 60):
    from openpyxl.utils import get_column_letter
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            if isinstance(v, (datetime, date)):
                s = v.isoformat(sep=" ") if isinstance(v, datetime) else v.isoformat()
            else:
                s = str(v)
            max_len = max(max_len, len(s))
        width = min(max_width, max(10, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_header(ws, header: List[str], row_idx: int = 1):
    from openpyxl.styles import Font
    for j, h in enumerate(header, start=1):
        cell = ws.cell(row=row_idx, column=j, value=h)
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def append_rows(ws, header: List[str], body_rows: List[List[Any]]):
    start = ws.max_row + 1
    for i, r in enumerate(body_rows):
        for j, v in enumerate(r, start=1):
            ws.cell(row=start + i, column=j, value=coerce_cell(v))


def main():
    args = parse_args()
    ensure_openpyxl()
    from openpyxl import Workbook, load_workbook

    text = read_text(args.input)

    if args.format == "json":
        dict_rows = parse_json(text)
        header = union_keys_order(dict_rows)
        body = [[r.get(k) for k in header] for r in dict_rows]
    elif args.format == "csv":
        header, body = parse_csv_text(text)
    else:
        header, body = parse_md_table(text)

    out = args.out
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)

    wb = None
    if args.mode == "overwrite":
        wb = Workbook()
        # remove default sheets then create our own with exact name
        while wb.sheetnames:
            wb.remove(wb[wb.sheetnames[0]])
        ws = wb.create_sheet(args.sheet)
        write_header(ws, header, 1)
        append_rows(ws, header, body)
    else:
        if os.path.exists(out):
            wb = load_workbook(out)
        else:
            wb = Workbook()
        if args.mode == "replace_sheet" and args.sheet in wb.sheetnames:
            wb.remove(wb[args.sheet])
        if args.sheet in wb.sheetnames:
            ws = wb[args.sheet]
            if args.mode == "append":
                # if empty sheet, add header
                if ws.max_row == 1 and all(ws.cell(1, j).value is None for j in range(1, max(1, len(header)) + 1)):
                    write_header(ws, header, 1)
                append_rows(ws, header, body)
            else:
                # replace_sheet handled by removal above
                pass
        else:
            ws = wb.create_sheet(args.sheet)
            write_header(ws, header, 1)
            append_rows(ws, header, body)

    # refresh autofilter range
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)

    wb.save(out)

    rows_written = len(body)
    cols_written = len(header)
    print(json.dumps({
        "path": os.path.abspath(out),
        "sheet": args.sheet,
        "mode": args.mode,
        "rows": rows_written,
        "cols": cols_written,
    }, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        eprint(f"ERROR: {e}")
        sys.exit(1)
